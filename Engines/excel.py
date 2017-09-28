#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import xlrd
import openpyxl
from openpyxl.styles import Border, Side
from xlutils.copy import copy

def check_excel_version(file):
    if file.endswith('.xlsx'):
        return 'XLSX'
    elif file.endswith('.xls'):
        return 'XLS'

def load_cases(file):
    file_type = check_excel_version(file)
    excel = eval(file_type)(file)
    excel.read_sheet()
    return excel.cases


class XLSX:
    def __init__(self, file):
        self.cases = []
        self.valid_rows = []
        self.file = file
        self.excel = openpyxl.load_workbook(self.file)

    # def read_sheet(self):
    def read_sheet(self):
        self.sheets = self.excel.worksheets
        for sheet in self.sheets:
            self.read_line(sheet)
        self.excel.save(self.file)

    def read_line(self, sheet):
        run_flag = None
        for i in range(2, sheet.max_row + 1):
            case_id = sheet.cell(row=i, column=2).value

            if case_id:
                # if case_id, then top border
                border = Border(top=Side(border_style='thin', color="FF000000"))
                # Y, case to run
                if sheet.cell(row=i, column=1).value.lower() == 'y':
                    run_flag = 1
                # N, case not to run
                else:
                    run_flag = 0
                    continue
            else:
                # no case_id, no top border
                border = Border(top=Side(border_style=None, color=None))
                # no case_id, and not belong to a case that is to run
                if run_flag == 0:
                    continue

            # set boarders for all cells
            for x in range(1, 13):
                sheet.cell(row=i, column=x).border = border

            # append the row to valid_rows, and add file/file_name/sheet_name/row info
            self.valid_rows.append([sheet.cell(row=i, column=x).value for x in range(1, 10)])
            self.valid_rows[-1].extend((self.file, os.path.basename(self.file), sheet.title, i))

            if i == sheet.max_row or sheet.cell(row=i + 1, column=2).value:
                self.cases.append(self.valid_rows)
                self.valid_rows = []

    def write_cell(self, sheet, row, column, value):
        self.excel.get_sheet_by_name(sheet).cell(row=row, column=column).value = value
        self.excel.save(self.file)

    def write_cells(self, results):
        for sheet, row, column, value in results:
            self.excel.get_sheet_by_name(sheet).cell(row=row, column=column).value = value
        self.excel.save(self.file)


# deprecated
class XLS:
    def __init__(self, file):
        self.cases = []
        self.valid_rows = []
        self.file = file
        self.excel = xlrd.open_workbook(self.file, formatting_info=True)
        self.wb = copy(self.excel)

    # def read_sheet(self):
    def read_sheet(self):
        self.sheets = self.excel.sheets()
        for i in self.sheets:
            self.read_line(i)

    def read_line(self, sheet):
        run_flag = None
        for i in range(1, sheet.nrows):
            case_id = sheet.cell(i, 1).value

            if case_id:
                # Y, new case
                if sheet.cell(i, 0).value.lower() == 'y':
                    run_flag = 1
                # Ignore N
                else:
                    run_flag = 0
                    continue
            # Ignore blank lines without run_flag 1
            elif run_flag == 0:
                continue

            self.valid_rows.append([sheet.cell(i, x).value for x in range(0, 9)])
            self.valid_rows[-1].extend((self.file, os.path.basename(self.file), sheet.name, i))

            if i == sheet.nrows - 1 or sheet.cell(i + 1, 1).value:
                self.cases.append(self.valid_rows)
                self.valid_rows = []

        def write_cell(self, sheet, row, column, value):
            sheet_index = self.excel.sheet_names().index(sheet)
            self.wb.get_sheet(sheet_index).write(row, column - 1, value)
            self.wb.save(self.file)

        def write_cells(self, results):
            for sheet, row, column, value in results:
                sheet_index = self.excel.sheet_names().index(sheet)
                self.wb.get_sheet(sheet_index).write(row, column - 1, value)
            self.wb.save(self.file)
