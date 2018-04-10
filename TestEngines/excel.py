#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import openpyxl
from openpyxl.styles import Border, Side


def write_excel(file, result):
    excel = openpyxl.load_workbook(file)
    for sheet, row, column, value in result:
        excel[sheet].cell(row=row, column=column).value = value
    excel.save(file)


class ReadAndFormatExcel:
    def __init__(self, file):
        self.cases = []
        self.logics = {}
        self.valid_rows = []
        self.file = file
        self.excel = openpyxl.load_workbook(self.file)
        self.sheets = self.excel.worksheets
        for sheet in self.sheets:
            self.read_line(sheet)
        self.excel.save(self.file)

    def read_line(self, sheet):
        run_flag = None
        for i in range(2, sheet.max_row + 1):
            case_id = sheet.cell(row=i, column=2).value

            if case_id:
                # if case_id, then top border
                border = Border(
                    top=Side(border_style='thin', color="FF000000")
                )
                # Y, case to run
                run_flag = sheet.cell(row=i, column=1).value.lower() or 'n'
                # N, case not to run
                if run_flag == 'n':
                    continue
            else:
                # no case_id, no top border
                border = Border(top=Side(border_style=None, color=None))
                # no case_id, and not belong to a case that is to run
                if run_flag == 'n':
                    continue

            # set boarders for all cells
            for x in range(1, 13):
                sheet.cell(row=i, column=x).border = border

            # append the row to valid_rows and
            # add file/file_name/sheet_name/row info
            self.valid_rows.append(
                [sheet.cell(row=i, column=x).value for x in range(1, 10)]
            )
            self.valid_rows[-1].extend(
                (self.file, os.path.basename(self.file), sheet.title, i)
            )

            if i == sheet.max_row or sheet.cell(row=i + 1, column=2).value:
                if run_flag == 'y':
                    self.cases.append(self.valid_rows)
                else:
                    self.logics[self.valid_rows[0][1]] = self.valid_rows
                self.valid_rows = []
